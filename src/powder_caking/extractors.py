from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True)
class BlockSpec:
    sigma1_kpa: float
    start_col: int
    end_col: int
    start_row: int
    end_row: int


BLOCK_SPECS = (
    BlockSpec(sigma1_kpa=20.0, start_col=12, end_col=14, start_row=9, end_row=28),
    BlockSpec(sigma1_kpa=11.0, start_col=17, end_col=19, start_row=9, end_row=24),
    BlockSpec(sigma1_kpa=3.1, start_col=22, end_col=24, start_row=9, end_row=16),
)


KINETICS_ROW_SPECS = (
    {"row": 8, "sigma1_kpa": 20.0},
    {"row": 9, "sigma1_kpa": 20.0},
    {"row": 10, "sigma1_kpa": 20.0},
    {"row": 11, "sigma1_kpa": 20.0},
    {"row": 18, "sigma1_kpa": 11.0},
    {"row": 19, "sigma1_kpa": 11.0},
    {"row": 20, "sigma1_kpa": 11.0},
    {"row": 27, "sigma1_kpa": 3.1},
    {"row": 28, "sigma1_kpa": 3.1},
)


def extract_mmp1_time_consolidation(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["MMP 1"]

    records: list[dict[str, object]] = []
    for spec in BLOCK_SPECS:
        current_temperature_c: float | None = None
        for row_idx in range(spec.start_row, spec.end_row + 1):
            label = ws.cell(row=row_idx, column=spec.start_col - 1).value
            time_h = ws.cell(row=row_idx, column=spec.start_col).value
            fc_pa = ws.cell(row=row_idx, column=spec.start_col + 1).value
            fc_err_pa = ws.cell(row=row_idx, column=spec.start_col + 2).value

            if isinstance(label, str) and "°C" in label:
                current_temperature_c = _parse_temperature(label)

            if not _is_number(time_h) or not _is_number(fc_pa):
                continue

            records.append(
                {
                    "material": "MMP 1",
                    "sigma1_kpa": spec.sigma1_kpa,
                    "temperature_c": current_temperature_c,
                    "time_h": float(time_h),
                    "fc_pa": float(fc_pa),
                    "fc_err_pa": float(fc_err_pa) if _is_number(fc_err_pa) else None,
                    "source_workbook": workbook_path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                    "source_range": f"{ws.cell(row=row_idx, column=spec.start_col).coordinate}:{ws.cell(row=row_idx, column=spec.start_col + 2).coordinate}",
                }
            )

    frame = pd.DataFrame.from_records(records)
    frame = frame.sort_values(["sigma1_kpa", "temperature_c", "time_h"], kind="stable").reset_index(drop=True)
    return frame


def extract_mmp1_kinetics_summary(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["MMP 1"]

    records: list[dict[str, object]] = []
    for spec in KINETICS_ROW_SPECS:
        row_idx = spec["row"]
        temperature_c = ws.cell(row=row_idx, column=1).value
        t_minus_tg_c = ws.cell(row=row_idx, column=2).value
        t_minus_tg_const_c = ws.cell(row=row_idx, column=3).value
        dfc_dt_pa_per_h = ws.cell(row=row_idx, column=4).value
        caking_time_40kpa_h = ws.cell(row=row_idx, column=5).value
        caking_time_20kpa_h = ws.cell(row=row_idx, column=6).value
        caking_time_critical_h = ws.cell(row=row_idx, column=9).value

        if not _is_number(temperature_c):
            continue

        records.append(
            {
                "material": "MMP 1",
                "sigma1_kpa": float(spec["sigma1_kpa"]),
                "temperature_c": float(temperature_c),
                "t_minus_tg_c": float(t_minus_tg_c) if _is_number(t_minus_tg_c) else None,
                "t_minus_tg_const_c": float(t_minus_tg_const_c) if _is_number(t_minus_tg_const_c) else None,
                "dfc_dt_pa_per_h": float(dfc_dt_pa_per_h) if _is_number(dfc_dt_pa_per_h) else None,
                "caking_time_40kpa_h": float(caking_time_40kpa_h) if _is_number(caking_time_40kpa_h) else None,
                "caking_time_20kpa_h": float(caking_time_20kpa_h) if _is_number(caking_time_20kpa_h) else None,
                "caking_time_critical_h": float(caking_time_critical_h) if _is_number(caking_time_critical_h) else None,
                "source_workbook": workbook_path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_range": f"A{row_idx}:I{row_idx}",
            }
        )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values(["sigma1_kpa", "temperature_c"], kind="stable").reset_index(drop=True)


def extract_relative_change_summary(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["realtive change"]

    records: list[dict[str, object]] = []
    for row_idx in range(4, 10):
        sigma1_kpa = ws.cell(row=row_idx, column=2).value
        moisture_wb_pct = ws.cell(row=row_idx, column=3).value
        moisture_db_pct = ws.cell(row=row_idx, column=4).value
        tg_c = ws.cell(row=row_idx, column=5).value
        a_param = ws.cell(row=row_idx, column=6).value
        k_param = ws.cell(row=row_idx, column=7).value
        temperature_c = ws.cell(row=row_idx, column=8).value
        t_minus_tg_c = ws.cell(row=row_idx, column=9).value
        caking_time_h = ws.cell(row=row_idx, column=10).value
        caking_time_d = ws.cell(row=row_idx, column=11).value
        caking_time_relative = ws.cell(row=row_idx, column=12).value

        if not _is_number(sigma1_kpa):
            continue

        records.append(
            {
                "material": "MMP 1",
                "sigma1_kpa": float(sigma1_kpa),
                "moisture_wb_pct": float(moisture_wb_pct) if _is_number(moisture_wb_pct) else None,
                "moisture_db_pct": float(moisture_db_pct) if _is_number(moisture_db_pct) else None,
                "tg_c": float(tg_c) if _is_number(tg_c) else None,
                "a_param": float(a_param) if _is_number(a_param) else None,
                "k_param": float(k_param) if _is_number(k_param) else None,
                "temperature_c": float(temperature_c) if _is_number(temperature_c) else None,
                "t_minus_tg_c": float(t_minus_tg_c) if _is_number(t_minus_tg_c) else None,
                "caking_time_h": float(caking_time_h) if _is_number(caking_time_h) else None,
                "caking_time_d": float(caking_time_d) if _is_number(caking_time_d) else None,
                "caking_time_relative": float(caking_time_relative) if _is_number(caking_time_relative) else None,
                "source_workbook": workbook_path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_range": f"B{row_idx}:L{row_idx}",
            }
        )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values(["sigma1_kpa", "moisture_db_pct"], kind="stable").reset_index(drop=True)


def extract_aw_tg_data(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["aw vs Tg"]

    records: list[dict[str, object]] = []
    for row_idx in range(5, 20):
        material = ws.cell(row=row_idx, column=1).value
        aw = ws.cell(row=row_idx, column=3).value
        tg_c = ws.cell(row=row_idx, column=4).value
        mdb_from_sheet = ws.cell(row=row_idx, column=5).value

        if not isinstance(material, str) or not _is_number(aw):
            continue

        records.append(
            {
                "material": material,
                "aw": float(aw),
                "tg_c": float(tg_c) if _is_number(tg_c) else None,
                "moisture_db_pct": float(mdb_from_sheet) if _is_number(mdb_from_sheet) else None,
                "source_workbook": workbook_path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_range": f"A{row_idx}:E{row_idx}",
            }
        )

    frame = pd.DataFrame.from_records(records)
    return frame.reset_index(drop=True)


def extract_critical_cake_strength(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["critical cake strength"]

    records: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        sigma_c_kpa = ws.cell(row=row_idx, column=2).value
        sieve_residue_pct = ws.cell(row=row_idx, column=3).value

        if not _is_number(sigma_c_kpa):
            continue

        records.append(
            {
                "material": "MMP 1",
                "sigma_c_kpa": float(sigma_c_kpa),
                "sieve_residue_pct": float(sieve_residue_pct) if _is_number(sieve_residue_pct) else None,
                "sieving_time_min": 1.0,
                "sieving_amplitude_mm": 1.0,
                "source_workbook": workbook_path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_range": f"B{row_idx}:C{row_idx}",
            }
        )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values("sigma_c_kpa", kind="stable").reset_index(drop=True)


def extract_permeation_time_series(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Permeationsmodell mit WDD est."]

    records: list[dict[str, object]] = []
    for row_idx in range(43, ws.max_row + 1):
        step = ws.cell(row=row_idx, column=1).value
        time_d = ws.cell(row=row_idx, column=2).value
        temperature_c = ws.cell(row=row_idx, column=10).value
        relative_humidity_pct = ws.cell(row=row_idx, column=11).value
        water_activity = ws.cell(row=row_idx, column=9).value
        tg_gordon_taylor_c = ws.cell(row=row_idx, column=14).value
        tg_linear_c = ws.cell(row=row_idx, column=15).value
        t_minus_tg_c = ws.cell(row=row_idx, column=16).value
        tg_mean_c = ws.cell(row=row_idx, column=17).value
        cake_strength_kpa = ws.cell(row=row_idx, column=18).value
        cumulative_water_kg = ws.cell(row=row_idx, column=19).value
        moisture_fraction = ws.cell(row=row_idx, column=4).value
        moisture_db_pct = ws.cell(row=row_idx, column=23).value

        if not _is_number(step) or not _is_number(time_d):
            continue

        records.append(
            {
                "step": int(step),
                "time_d": float(time_d),
                "temperature_c": float(temperature_c) if _is_number(temperature_c) else None,
                "relative_humidity_pct": float(relative_humidity_pct) if _is_number(relative_humidity_pct) else None,
                "water_activity": float(water_activity) if _is_number(water_activity) else None,
                "tg_gordon_taylor_c": float(tg_gordon_taylor_c) if _is_number(tg_gordon_taylor_c) else None,
                "tg_linear_c": float(tg_linear_c) if _is_number(tg_linear_c) else None,
                "t_minus_tg_c": float(t_minus_tg_c) if _is_number(t_minus_tg_c) else None,
                "tg_mean_c": float(tg_mean_c) if _is_number(tg_mean_c) else None,
                "cake_strength_kpa": float(cake_strength_kpa) if _is_number(cake_strength_kpa) else None,
                "cumulative_water_kg": float(cumulative_water_kg) if _is_number(cumulative_water_kg) else None,
                "moisture_fraction": float(moisture_fraction) if _is_number(moisture_fraction) else None,
                "moisture_db_pct": float(moisture_db_pct) if _is_number(moisture_db_pct) else None,
                "source_workbook": workbook_path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_range": f"A{row_idx}:W{row_idx}",
            }
        )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values("step", kind="stable").reset_index(drop=True)


def extract_real_container_logger_profile(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Permeationsmodell mit WDD est."]

    records: list[dict[str, float]] = []
    for row_idx in range(43, ws.max_row + 1):
        time_d = ws.cell(row=row_idx, column=2).value
        temperature_c = ws.cell(row=row_idx, column=10).value
        relative_humidity_pct = ws.cell(row=row_idx, column=11).value

        if not _is_number(time_d):
            continue

        records.append(
            {
                "time_d": float(time_d),
                "temperature_c": float(temperature_c) if _is_number(temperature_c) else None,
                "relative_humidity_pct": float(relative_humidity_pct) if _is_number(relative_humidity_pct) else None,
            }
        )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values("time_d", kind="stable").reset_index(drop=True)


def extract_table2_scenario_series(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Tabelle2"]

    scenarios = (
        {"initial_moisture_db_pct": 3.8, "fc_col": 2, "moisture_col": 6, "tg_col": 10},
        {"initial_moisture_db_pct": 4.0, "fc_col": 3, "moisture_col": 7, "tg_col": 11},
        {"initial_moisture_db_pct": 4.2, "fc_col": 4, "moisture_col": 8, "tg_col": 12},
    )

    records: list[dict[str, object]] = []
    for row_idx in range(5, ws.max_row + 1):
        time_d = ws.cell(row=row_idx, column=2).value
        if not _is_number(time_d):
            continue

        for scenario in scenarios:
            fc_kpa = ws.cell(row=row_idx, column=scenario["fc_col"] + 1).value
            moisture_db_pct = ws.cell(row=row_idx, column=scenario["moisture_col"] + 1).value
            tg_c = ws.cell(row=row_idx, column=scenario["tg_col"] + 1).value

            records.append(
                {
                    "initial_moisture_db_pct": float(scenario["initial_moisture_db_pct"]),
                    "time_d": float(time_d),
                    "cake_strength_kpa": float(fc_kpa) if _is_number(fc_kpa) else None,
                    "moisture_db_pct": float(moisture_db_pct) if _is_number(moisture_db_pct) else None,
                    "tg_c": float(tg_c) if _is_number(tg_c) else None,
                    "source_workbook": workbook_path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                    "source_range": f"B{row_idx}:M{row_idx}",
                }
            )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values(["initial_moisture_db_pct", "time_d"], kind="stable").reset_index(drop=True)


def extract_wdd_permeability_summary(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Messung WDD Sackmaterial"]

    records: list[dict[str, object]] = []
    for row_idx in range(8, 11):
        temperature_c = ws.cell(row=row_idx, column=51).value
        mass_gain_g_per_d = ws.cell(row=row_idx, column=52).value
        wdd_g_per_m2_d = ws.cell(row=row_idx, column=53).value
        k_over_delta_kg_per_m2_d_pa = ws.cell(row=row_idx, column=54).value
        inverse_temperature_1_per_k = ws.cell(row=row_idx, column=55).value
        ln_k_over_delta = ws.cell(row=row_idx, column=56).value

        if not _is_number(temperature_c):
            continue

        records.append(
            {
                "temperature_c": float(temperature_c),
                "mass_gain_g_per_d": float(mass_gain_g_per_d) if _is_number(mass_gain_g_per_d) else None,
                "wdd_g_per_m2_d": float(wdd_g_per_m2_d) if _is_number(wdd_g_per_m2_d) else None,
                "k_over_delta_kg_per_m2_d_pa": float(k_over_delta_kg_per_m2_d_pa)
                if _is_number(k_over_delta_kg_per_m2_d_pa)
                else None,
                "inverse_temperature_1_per_k": float(inverse_temperature_1_per_k)
                if _is_number(inverse_temperature_1_per_k)
                else None,
                "ln_k_over_delta": float(ln_k_over_delta) if _is_number(ln_k_over_delta) else None,
                "source_workbook": workbook_path.name,
                "source_sheet": ws.title,
                "source_row": row_idx,
                "source_range": f"AY{row_idx}:BD{row_idx}",
            }
        )

    return pd.DataFrame.from_records(records).sort_values("temperature_c").reset_index(drop=True)


def extract_wdd_arrhenius_parameters(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Messung WDD Sackmaterial"]

    values = [
        {
            "parameter": "gas_constant",
            "symbol": "R",
            "value": ws.cell(row=7, column=59).value,
            "unit": "J/(kmol*K)",
            "source_cell": "BG7",
        },
        {
            "parameter": "negative_ea_over_r",
            "symbol": "-Ea/R",
            "value": ws.cell(row=8, column=59).value,
            "unit": "K",
            "source_cell": "BG8",
        },
        {
            "parameter": "ln_k0",
            "symbol": "ln(k0)",
            "value": ws.cell(row=9, column=59).value,
            "unit": "-",
            "source_cell": "BG9",
        },
        {
            "parameter": "activation_energy",
            "symbol": "Ea",
            "value": ws.cell(row=11, column=59).value,
            "unit": "J/kmol",
            "source_cell": "BG11",
        },
        {
            "parameter": "pre_exponential_factor",
            "symbol": "k0",
            "value": ws.cell(row=12, column=59).value,
            "unit": "kg/(m2*d*Pa)",
            "source_cell": "BG12",
        },
    ]

    for item in values:
        item["source_workbook"] = workbook_path.name
        item["source_sheet"] = ws.title

    return pd.DataFrame.from_records(values)


def extract_wdd_measurement_timeseries(workbook_path: str | Path) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Messung WDD Sackmaterial"]

    blocks = (
        {"temperature_c": 23.0, "mass_col": 6, "time_col": 9, "delta_mass_col": 10, "time_d_col": 11, "start_row": 11},
        {"temperature_c": 35.0, "mass_col": 26, "time_col": 29, "delta_mass_col": 30, "time_d_col": 31, "start_row": 11},
        {"temperature_c": 15.0, "mass_col": 39, "time_col": 41, "delta_mass_col": 42, "time_d_col": 43, "start_row": 6},
    )

    records: list[dict[str, object]] = []
    for block in blocks:
        for row_idx in range(block["start_row"], ws.max_row + 1):
            mass_g = ws.cell(row=row_idx, column=block["mass_col"]).value
            elapsed_min = ws.cell(row=row_idx, column=block["time_col"]).value
            delta_mass_g = ws.cell(row=row_idx, column=block["delta_mass_col"]).value
            time_d = ws.cell(row=row_idx, column=block["time_d_col"]).value

            if not _is_number(mass_g) and not _is_number(time_d):
                continue
            if _is_number(time_d) and float(time_d) > 200:
                continue

            records.append(
                {
                    "temperature_c": block["temperature_c"],
                    "mass_g": float(mass_g) if _is_number(mass_g) else None,
                    "elapsed_min": float(elapsed_min) if _is_number(elapsed_min) else None,
                    "delta_mass_g": float(delta_mass_g) if _is_number(delta_mass_g) else None,
                    "time_d": float(time_d) if _is_number(time_d) else None,
                    "source_workbook": workbook_path.name,
                    "source_sheet": ws.title,
                    "source_row": row_idx,
                }
            )

    frame = pd.DataFrame.from_records(records)
    return frame.sort_values(["temperature_c", "time_d"], kind="stable").reset_index(drop=True)


def _parse_temperature(label: str) -> float:
    return float(label.replace("°C", "").replace(",", ".").strip())


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not pd.isna(value)
